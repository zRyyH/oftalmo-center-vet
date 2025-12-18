from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def criar_estilos():
    borda = Side(style='thin', color='000000')
    return {
        'header_font': Font(bold=True, color='FFFFFF', size=11),
        'header_fill': PatternFill('solid', fgColor='2E7D32'),
        'cell_font': Font(size=10),
        'alt_fill': PatternFill('solid', fgColor='E8F5E9'),
        'borda': Border(left=borda, right=borda, top=borda, bottom=borda),
        'centro': Alignment(horizontal='center', vertical='center'),
        'sim_fill': PatternFill('solid', fgColor='4CAF50'),
        'nao_fill': PatternFill('solid', fgColor='F44336'),
    }


def aplicar_estilo_header(ws, estilos, colunas):
    for col, nome in enumerate(colunas, 1):
        celula = ws.cell(row=1, column=col, value=nome)
        celula.font = estilos['header_font']
        celula.fill = estilos['header_fill']
        celula.alignment = estilos['centro']
        celula.border = estilos['borda']


def aplicar_estilo_linha(ws, linha, estilos, num_cols, conciliado, zebra=False):
    for col in range(1, num_cols + 1):
        celula = ws.cell(row=linha, column=col)
        celula.font = estilos['cell_font']
        celula.border = estilos['borda']
        celula.alignment = estilos['centro']
        if col == 5:
            celula.fill = estilos['sim_fill'] if conciliado else estilos['nao_fill']
        elif zebra:
            celula.fill = estilos['alt_fill']


def formatar_data(data_str):
    return data_str[:10] if data_str else ''


def gerar_excel(conciliations, arquivo='conciliacoes.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Conciliações'
    
    colunas = ['Data', 'Valor Sicoob', 'Soma Merchant', 'Descrição', 'Conciliado', 'Diferença']
    larguras = [12, 15, 15, 35, 12, 12]
    estilos = criar_estilos()
    
    aplicar_estilo_header(ws, estilos, colunas)
    
    for i, item in enumerate(conciliations, 2):
        ws.cell(row=i, column=1, value=formatar_data(item['data']))
        ws.cell(row=i, column=2, value=item['valor_sicoob'])
        ws.cell(row=i, column=3, value=item['soma_merchant'])
        ws.cell(row=i, column=4, value=item['desc_inf_complementar'])
        ws.cell(row=i, column=5, value='Sim' if item['conciliado'] else 'Não')
        ws.cell(row=i, column=6, value=item['diferenca'])
        
        aplicar_estilo_linha(ws, i, estilos, len(colunas), item['conciliado'], zebra=(i % 2 == 0))
    
    for col, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = largura
    
    ws.freeze_panes = 'A2'
    wb.save(arquivo)
    return arquivo


if __name__ == '__main__':
    dados = [
        {"valor_sicoob": 298.35, "data": "2025-12-02 06:46:00.000Z", "soma_merchant": 298.35, "desc_inf_complementar": "EVOLUSERVICES _Deb._Maestro", "conciliado": True, "diferenca": 0.0},
        {"valor_sicoob": 316.48, "data": "2025-12-03 07:07:00.000Z", "soma_merchant": 316.48, "desc_inf_complementar": "EVOLUSERVICES _Deb._Maestro", "conciliado": True, "diferenca": 0.0},
        {"valor_sicoob": 499.54, "data": "2025-12-04 07:09:00.000Z", "soma_merchant": 499.54, "desc_inf_complementar": "EVOLUSERVICES _Cred._Visa", "conciliado": True, "diferenca": 0.0},
        {"valor_sicoob": 495.44, "data": "2025-12-05 07:08:00.000Z", "soma_merchant": 495.44, "desc_inf_complementar": "EVOLUSERVICES _Cred._Mastercard", "conciliado": True, "diferenca": 0.0},
        {"valor_sicoob": 434.77, "data": "2025-12-05 07:08:00.000Z", "soma_merchant": 434.77, "desc_inf_complementar": "EVOLUSERVICES _Cred._Visa", "conciliado": True, "diferenca": 0.0},
        {"valor_sicoob": 370.88, "data": "2025-12-08 07:07:00.000Z", "soma_merchant": 370.88, "desc_inf_complementar": "EVOLUSERVICES _Deb._Maestro", "conciliado": True, "diferenca": 0.0},
        {"valor_sicoob": 907.08, "data": "2025-12-08 07:08:00.000Z", "soma_merchant": 907.08, "desc_inf_complementar": "EVOLUSERVICES _Cred._Visa", "conciliado": True, "diferenca": 0.0},
    ]
    
    gerar_excel(dados, '/mnt/user-data/outputs/conciliacoes.xlsx')