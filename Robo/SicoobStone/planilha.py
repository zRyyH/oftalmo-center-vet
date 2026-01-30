from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from openpyxl import Workbook


COR_HEADER_BG = "2E7D32"
COR_HEADER_FONT = "FFFFFF"
COR_ZEBRA = "E8F5E9"
COR_BRANCO = "FFFFFF"
COR_MATCH = "C8E6C9"
COR_NAO_MATCH = "FFCDD2"
COR_FONTE = "212121"
COR_TOTAL_BG = "37474F"
COR_TOTAL_FONT = "FFFFFF"

BORDA = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


def ajustar_largura_colunas(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ""
                max_length = max(max_length, len(val))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 3


def gerar_planilha_sicoob(dados, caminho_saida="sicoob_stone.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)

    registros = []
    for r in dados.get("conciliados", []):
        registros.append(r)
    for r in dados.get("nao_conciliados", []):
        registros.append(r)

    por_mes = defaultdict(list)
    for r in registros:
        data_ref = r.get("data_sicoob") or r.get("data_stone") or ""
        if data_ref:
            mes = data_ref[5:7] + "-" + data_ref[:4]
            por_mes[mes].append(r)

    meses_ordenados = sorted(por_mes.keys(), key=lambda x: x[3:7] + x[:2])

    cabecalhos = [
        "Data Sicoob",
        "Data Stone",
        "Conciliado",
        "Valor Sicoob (R$)",
        "Valor Stone (R$)",
        "Descrição Sicoob",
        "Info Complementar",
        "Bandeira",
    ]
    campos = [
        "data_sicoob",
        "data_stone",
        "conciliado",
        "valor_sicoob",
        "valor_stone",
        "descricao",
        "desc_inf_complementar",
        "bandeira",
    ]

    for mes in meses_ordenados:
        ws = wb.create_sheet(title=mes)
        regs = sorted(
            por_mes[mes],
            key=lambda x: x.get("data_sicoob") or x.get("data_stone") or "",
            reverse=True,
        )
        _criar_aba(ws, cabecalhos, regs, campos)

    wb.save(caminho_saida)
    return caminho_saida


def _criar_aba(ws, cabecalhos, registros, campos):
    header_font = Font(bold=True, size=11, color=COR_HEADER_FONT)
    header_fill = PatternFill("solid", fgColor=COR_HEADER_BG)
    alinhamento = Alignment(horizontal="center", vertical="center")

    for col, texto in enumerate(cabecalhos, 1):
        cel = ws.cell(row=1, column=col, value=texto)
        cel.font = header_font
        cel.fill = header_fill
        cel.alignment = alinhamento
        cel.border = BORDA

    data_font = Font(size=10, color=COR_FONTE)
    fill_zebra = PatternFill("solid", fgColor=COR_ZEBRA)
    fill_branco = PatternFill("solid", fgColor=COR_BRANCO)
    fill_match = PatternFill("solid", fgColor=COR_MATCH)
    fill_nao = PatternFill("solid", fgColor=COR_NAO_MATCH)

    total_sicoob = 0
    total_stone = 0

    for row_idx, registro in enumerate(registros, 2):
        conciliado = registro.get("conciliado", "")
        linha_par = row_idx % 2 == 0

        for col_idx, campo in enumerate(campos, 1):
            valor = registro.get(campo)
            cel = ws.cell(row=row_idx, column=col_idx, value=valor)
            cel.font = data_font
            cel.alignment = alinhamento
            cel.border = BORDA

            if campo == "conciliado":
                cel.fill = fill_match if conciliado == "SIM" else fill_nao
            else:
                cel.fill = fill_zebra if linha_par else fill_branco

            if campo in ["valor_sicoob", "valor_stone"]:
                cel.number_format = "R$ #,##0.00"

        total_sicoob += registro.get("valor_sicoob", 0) or 0
        total_stone += registro.get("valor_stone", 0) or 0

    if registros:
        row_total = len(registros) + 2
        total_font = Font(bold=True, size=11, color=COR_TOTAL_FONT)
        total_fill = PatternFill("solid", fgColor=COR_TOTAL_BG)

        for col in range(1, len(cabecalhos) + 1):
            cel = ws.cell(row=row_total, column=col)
            cel.font = total_font
            cel.fill = total_fill
            cel.border = BORDA
            cel.alignment = alinhamento

        ws.cell(row=row_total, column=1, value="TOTAL")
        ws.cell(row=row_total, column=4, value=total_sicoob).number_format = (
            "R$ #,##0.00"
        )
        ws.cell(row=row_total, column=5, value=total_stone).number_format = (
            "R$ #,##0.00"
        )

    ajustar_largura_colunas(ws)

    ws.freeze_panes = "A2"

    if registros:
        ultima_col = get_column_letter(len(cabecalhos))
        ultima_linha = len(registros) + 1
        ws.auto_filter.ref = f"A1:{ultima_col}{ultima_linha}"
