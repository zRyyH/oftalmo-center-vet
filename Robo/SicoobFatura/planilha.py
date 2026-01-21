from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict


COR_CABECALHO = PatternFill("solid", fgColor="90EE90")
COR_ZEBRA = PatternFill("solid", fgColor="E8F5E9")
COR_STATUS_OK = PatternFill("solid", fgColor="C6EFCE")
COR_STATUS_ERRO = PatternFill("solid", fgColor="FFC7CE")
BORDA = Border(
    left=Side(style="thin", color="808080"),
    right=Side(style="thin", color="808080"),
    top=Side(style="thin", color="808080"),
    bottom=Side(style="thin", color="808080"),
)

COLUNAS = ["Data", "Descrição", "Fornecedor", "Valor", "Parcela", "Vencimento"]


def extrair_mes_ano(data_str):
    try:
        if "T" in data_str or "Z" in data_str:
            data_str = data_str[:10]
        dt = datetime.strptime(data_str[:10], "%Y-%m-%d")
        return dt.strftime("%m-%Y")
    except:
        return None


def formatar_data(data_str):
    try:
        if data_str and len(data_str) >= 10:
            return datetime.strptime(data_str[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except:
        pass
    return data_str or ""


def formatar_valor(valor):
    if valor is None:
        return 0
    return abs(float(valor))


def aplicar_estilo_cabecalho(ws):
    for col, titulo in enumerate(COLUNAS, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = Font(bold=True, size=11)
        cell.fill = COR_CABECALHO
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDA


def ajustar_larguras(ws):
    for col in range(1, len(COLUNAS) + 1):
        max_len = len(COLUNAS[col - 1])
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 50)


def gerar_planilha_sicoob(dados, caminho_saida="Relatorios/Cartão De Credito.xlsx"):
    registros_por_mes = defaultdict(
        lambda: {"despesas": [], "valor_fatura": 0, "total_despesas": 0}
    )

    for item in dados:
        data_fatura = item.get("data_fatura", "")
        mes_ano = extrair_mes_ano(data_fatura) or "Sem Data"

        registros_por_mes[mes_ano]["valor_fatura"] = item.get("valor_fatura", 0)
        registros_por_mes[mes_ano]["total_despesas"] = item.get("total_despesas", 0)

        for desp in item.get("despesas", []):
            registro = {
                "data": formatar_data(desp.get("data", data_fatura)),
                "descricao": desp.get("descricao", ""),
                "fornecedor": desp.get("fornecedor", "").replace(
                    "Cartão de crédito", ""
                ),
                "valor": formatar_valor(desp.get("valor")),
                "parcela": desp.get("parcela", ""),
                "vencimento": formatar_data(desp.get("vencimento", "")),
                "data_sort": desp.get("data", data_fatura),
            }
            registros_por_mes[mes_ano]["despesas"].append(registro)

    wb = Workbook()
    wb.remove(wb.active)

    for mes_ano in sorted(registros_por_mes.keys(), reverse=True):
        dados_mes = registros_por_mes[mes_ano]
        despesas = dados_mes["despesas"]
        despesas.sort(key=lambda x: x["data_sort"], reverse=True)

        ws = wb.create_sheet(title=mes_ano[:31])
        aplicar_estilo_cabecalho(ws)

        linha_atual = 2
        for reg in despesas:
            valores = [
                reg["data"],
                reg["descricao"],
                reg["fornecedor"],
                reg["valor"],
                reg["parcela"],
                reg["vencimento"],
            ]

            for col, valor in enumerate(valores, 1):
                cell = ws.cell(row=linha_atual, column=col, value=valor)
                cell.border = BORDA
                cell.font = Font(size=10)

                if col in [1, 4, 5, 6]:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")

                if linha_atual % 2 == 0:
                    cell.fill = COR_ZEBRA

            ws.cell(row=linha_atual, column=4).number_format = "R$ #,##0.00"
            linha_atual += 1

        # Resumo no final
        linha_atual += 1
        valor_fatura = dados_mes["valor_fatura"]
        total_despesas = dados_mes["total_despesas"]
        diferenca = round(valor_fatura - total_despesas, 2)
        tem_erro = abs(diferenca) > 0.01

        # Total Despesas
        cell_label = ws.cell(row=linha_atual, column=3, value="TOTAL DESPESAS:")
        cell_label.font = Font(bold=True, size=14)
        cell_label.alignment = Alignment(horizontal="right")

        cell_valor = ws.cell(row=linha_atual, column=4, value=total_despesas)
        cell_valor.font = Font(bold=True, size=14)
        cell_valor.number_format = "R$ #,##0.00"
        cell_valor.alignment = Alignment(horizontal="center")

        linha_atual += 1

        # Valor Fatura
        cell_label = ws.cell(row=linha_atual, column=3, value="VALOR FATURA:")
        cell_label.font = Font(bold=True, size=14)
        cell_label.alignment = Alignment(horizontal="right")

        cell_valor = ws.cell(row=linha_atual, column=4, value=valor_fatura)
        cell_valor.font = Font(bold=True, size=14)
        cell_valor.number_format = "R$ #,##0.00"
        cell_valor.alignment = Alignment(horizontal="center")

        linha_atual += 1

        # Diferença
        cell_label = ws.cell(row=linha_atual, column=3, value="DIFERENÇA:")
        cell_label.font = Font(
            bold=True, size=14, color="CC0000" if tem_erro else "006400"
        )
        cell_label.alignment = Alignment(horizontal="right")

        cell_valor = ws.cell(row=linha_atual, column=4, value=diferenca)
        cell_valor.font = Font(
            bold=True, size=14, color="CC0000" if tem_erro else "006400"
        )
        cell_valor.number_format = "R$ #,##0.00"
        cell_valor.alignment = Alignment(horizontal="center")
        cell_valor.fill = COR_STATUS_ERRO if tem_erro else COR_STATUS_OK

        ajustar_larguras(ws)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:F{linha_atual - 4}"

    wb.save(caminho_saida)
    return caminho_saida
